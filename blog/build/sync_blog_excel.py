import hashlib
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
BLOG_DATA_DIR = REPO_ROOT / 'blog' / 'blog-data'
EXCEL_PATH = REPO_ROOT / 'blog' / 'blog-data-summary.xlsx'


def parse_frontmatter(text: str) -> Dict[str, str]:
    match = re.match(r'^---\s*\n([\s\S]*?)\n---\s*\n?', text)
    if not match:
        return {}

    frontmatter = {}
    for line in match.group(1).splitlines():
        stripped = line.strip()
        if not stripped or stripped.startswith('#'):
            continue
        if ':' not in line:
            continue
        key, value = line.split(':', 1)
        key = key.strip()
        value = value.strip()
        if key:
            frontmatter[key] = value
    return frontmatter


def format_scalar(value: object) -> str:
    if isinstance(value, bool):
        return 'true' if value else 'false'
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).strip()
    if not text:
        return ''
    if re.search(r'[:#\[\]\{\},&*?!|@`\\]', text) or text.startswith(('"', "'")) or text.startswith('-'):
        return json_quote(text)
    return text


def json_quote(text: str) -> str:
    import json
    return json.dumps(text)


def parse_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {'1', 'true', 'yes', 'y', 'on'}


def derive_posted(frontmatter: Dict[str, str]) -> str:
    posted_value = frontmatter.get('posted')
    if posted_value is not None:
        return 'Yes' if parse_boolean(posted_value) else 'No'

    draft = parse_boolean(frontmatter.get('draft', False))
    if draft:
        return 'No'

    date_value = frontmatter.get('date', '').strip()
    if not date_value:
        return 'Yes'

    try:
        published_date = datetime.strptime(date_value, '%Y-%m-%d').date()
    except ValueError:
        return 'Yes'

    return 'Yes' if published_date <= date.today() else 'No'


def update_markdown_frontmatter(file_path: Path, values: Dict[str, str]) -> None:
    return


def build_rows() -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for markdown_path in sorted(BLOG_DATA_DIR.glob('*.md')):
        text = markdown_path.read_text(encoding='utf-8')
        frontmatter = parse_frontmatter(text)
        title = frontmatter.get('title', markdown_path.stem.replace('-', ' ').title())
        category = frontmatter.get('category', '')
        date_value = frontmatter.get('date', '')
        posted = derive_posted(frontmatter)

        rows.append({
            'file_name': markdown_path.name,
            'title': title,
            'category': category,
            'date': date_value,
            'posted': posted,
        })
    return rows


def write_excel(rows: List[Dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'blog-data'
    headers = ['file_name', 'title', 'category', 'date', 'posted']
    sheet.append(headers)
    for row in rows:
        sheet.append([row['file_name'], row['title'], row['category'], row['date'], row['posted']])
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 60)
    workbook.save(EXCEL_PATH)


def read_excel_rows() -> List[Dict[str, str]]:
    if not EXCEL_PATH.exists():
        return []
    workbook = load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    rows: List[Dict[str, str]] = []
    for raw_row in sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, raw_row))
        if not record or not record.get('file_name'):
            continue
        rows.append({
            'file_name': str(record.get('file_name', '')),
            'title': str(record.get('title', '')),
            'category': str(record.get('category', '')),
            'date': str(record.get('date', '')),
            'posted': str(record.get('posted', 'Yes')),
        })
    return rows


def apply_excel_to_markdown(rows: List[Dict[str, str]]) -> None:
    return


def snapshot_state() -> Dict[str, object]:
    markdown_state = {}
    for markdown_path in sorted(BLOG_DATA_DIR.glob('*.md')):
        file_bytes = markdown_path.read_bytes()
        markdown_state[markdown_path.name] = {
            'hash': hashlib.sha256(file_bytes).hexdigest(),
            'size': len(file_bytes),
        }

    excel_bytes = EXCEL_PATH.read_bytes() if EXCEL_PATH.exists() else b''
    return {
        'markdown': markdown_state,
        'excel_hash': hashlib.sha256(excel_bytes).hexdigest(),
    }


def watch_for_changes(interval: float = 2.0) -> None:
    last_state = None
    while True:
        try:
            state = snapshot_state()
            if last_state is None:
                write_excel(build_rows())
                print(f'Initial sync complete: {EXCEL_PATH}')
                last_state = state
                time.sleep(interval)
                continue

            if state != last_state:
                markdown_changed = state['markdown'] != last_state['markdown']
                excel_changed = state['excel_hash'] != last_state['excel_hash']

                if markdown_changed:
                    write_excel(build_rows())
                    print('Blog data change detected; refreshed workbook')
                elif excel_changed:
                    write_excel(build_rows())
                    print('Excel change detected; refreshed workbook from markdown source')

                last_state = snapshot_state()
            time.sleep(interval)
        except KeyboardInterrupt:
            print('Stopped blog Excel watcher')
            break


def main() -> None:
    if '--apply-excel' in sys.argv:
        apply_excel_to_markdown(read_excel_rows())
        write_excel(build_rows())
        print(f'Synced Excel -> markdown and refreshed {EXCEL_PATH}')
        return

    if '--watch' in sys.argv:
        watch_for_changes()
        return

    rows = build_rows()
    write_excel(rows)
    print(f'Wrote {len(rows)} rows to {EXCEL_PATH}')


if __name__ == '__main__':
    main()
